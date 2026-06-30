"""Build export.xlsx — header row matches db.sql columns and db.js DB_COLUMNS exactly.
No sample data rows."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "expense_data"

headers = ["No.", "Date", "Time", "From", "To", "Amount", "Type", "Note", "Created At"]

# styling — white & black theme to match the app
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="000000")
center = Alignment(horizontal="center", vertical="center")
thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, title in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=title)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# column widths
widths = [6, 14, 12, 26, 26, 12, 10, 24, 22]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws.freeze_panes = "A2"
wb.save("export.xlsx")
print("export.xlsx written with headers:", headers)
